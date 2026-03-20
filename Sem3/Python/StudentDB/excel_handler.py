import os
from openpyxl import Workbook, load_workbook

FILE_NAME = "studentDB.xlsx"

def create_excel_if_not_exists():
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active

        ws.append([
            "Name", "Address", "DOB",
            "Department", "Class X %",
            "Class XII %", "Gender"
        ])

        wb.save(FILE_NAME)


def save_to_excel(data):
    create_excel_if_not_exists()

    wb = load_workbook(FILE_NAME)
    ws = wb.active

    ws.append(data)

    wb.save(FILE_NAME)


def open_excel_file():
    create_excel_if_not_exists()
    if os.path.exists(FILE_NAME):
        os.startfile(FILE_NAME)